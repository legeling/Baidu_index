#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：Achieve 
@File    ：简易版.py
@IDE     ：PyCharm 
@Author  ：lingxiaotian
@Date    ：2023/10/25 21:17 
'''

import requests
import json
import os
from datetime import datetime, timedelta
import openpyxl


# 解码函数
def decrypt(ptbk, index_data):
    n = len(ptbk)//2
    a = dict(zip(ptbk[:n], ptbk[n:]))
    return "".join([a[s] for s in index_data])

# 获取数据源并暂存至文件中
def get_index_data(keys,year):
    words = [[{"name": keys, "wordType": 1}]]
    words = str(words).replace(" ", "").replace("'", "\"")
    startDate = f"{year}-01-01"
    endDate = f"{year}-12-31"
    url = f'http://index.baidu.com/api/SearchApi/index?area=0&word={words}&startDate={startDate}&endDate={endDate}'
    # 请求头配置
    headers = {
        "Connection": "keep-alive",
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.198 Safari/537.36",
        "Sec-Fetch-Site": "same-origin",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Dest": "empty",
        "Cipher-Text": "1698156005330_1698238860769_ZPrC2QTaXriysBT+5sgXcnbTX3/lW65av4zgu9uR1usPy82bArEg4m9deebXm7/O5g6QWhRxEd9/r/hqHad2WnVFVVWybHPFg3YZUUCKMTIYFeSUIn23C6HdTT1SI8mxsG5mhO4X9nnD6NGI8hF8L5/G+a5cxq+b21PADOpt/XB5eu/pWxNdwfa12krVNuYI1E8uHQ7TFIYjCzLX9MoJzPU6prjkgJtbi3v0X7WGKDJw9hwnd5Op4muW0vWKMuo7pbxUNfEW8wPRmSQjIgW0z5p7GjNpsg98rc3FtHpuhG5JFU0kZ6tHgU8+j6ekZW7+JljdyHUMwEoBOh131bGl+oIHR8vw8Ijtg8UXr0xZqcZbMEagEBzWiiKkEAfibCui59hltAgW5LG8IOtBDqp8RJkbK+IL5GcFkNaXaZfNMpI=",
        "Referer": "https://index.baidu.com/v2/main/index.html",
        "Accept-Language": "zh-CN,zh;q=0.9",
        'Cookie': Cookie}
    res = requests.get(url, headers=headers)
    res_json = res.json()
    # print(res_json)
    if res_json["message"] == "bad request":
        print("抓取关键词："+keys+" 失败，请检查cookie或者关键词是否存在")
    else:
        # 获取特征值
        data = res_json['data']
        # print(data)
        uniqid = data["uniqid"]
        url = f'http://index.baidu.com/Interface/ptbk?uniqid={uniqid}'
        res = requests.get(url, headers=headers)
        # 获取解码字
        ptbk = res.json()['data']

        #创建暂存文件夹
        os.makedirs('res', exist_ok=True)
        filename = f"{keys}_{year}.json"
        file_path = os.path.join('res', filename)
        with open(file_path, 'w', encoding='utf-8') as json_file:
            json.dump(res_json, json_file, ensure_ascii=False, indent=4)
        return file_path,ptbk

def reCode(file_path,ptbk):
    # 读取暂存文件
    with open(file_path, 'r', encoding='utf-8') as file:
        res = json.load(file)
    data = res['data']
    li = data['userIndexes'][0]['all']['data']
    startDate = data['userIndexes'][0]['all']['startDate']
    year_str = startDate[:4]  # 使用切片取前四个字符，即年份部分
    try:
        # 将年份字符串转换为整数
        year = int(year_str)
        # 根据年份判断是否为闰年
        if (year % 4 == 0 and year % 100 != 0) or (year % 400 == 0):
            year = 366
        else:
            year = 365
    except :
        year =365

    if li =='':
        result = {}
        name = data['userIndexes'][0]['word'][0]['name']
        tep_all = []
        while len(tep_all) < year:
            tep_all.insert(0, 0)
        result["name"] = name
        result["data"] = tep_all
    else:
        ptbk = ptbk
        result = {}
        for userIndexe in data['userIndexes']:
            name = userIndexe['word'][0]['name']
            index_all = userIndexe['all']['data']
            try:
                index_all_data = [int(e) for e in decrypt(ptbk, index_all).split(",")]
                tmp_all = index_all_data
            except:
                tmp_all = []
            while len(tmp_all) < year:
                tmp_all.insert(0, 0)
            result["name"] = name
            result["data"] = tmp_all
    return result

#创建日期表格
def create_excel(start_year, end_year):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    # 设置第一行的标题
    sheet['A1'] = '日期'

    # 计算日期范围
    start_date = datetime(start_year, 1, 1)
    end_date = datetime(end_year, 12, 31)

    # 逐行填充日期
    current_date = start_date
    row = 2  # 从第二行开始
    while current_date <= end_date:
        sheet[f'A{row}'] = current_date.strftime('%Y-%m-%d')
        current_date += timedelta(days=1)
        row += 1

    # 保存 Excel 文件
    filename = f'百度指数数据-{start_year}-{end_year}.xlsx'
    workbook.save(filename)
    return filename

#为文件写入数据
def write_to_excel(file_name, name, data,i):
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_name)
        # 获取默认的工作表（第一个工作表）
        sheet = workbook.active
        # 将名称写入第一行第i列
        sheet.cell(row=1, column=i, value=name)
        # 将数据写入从第二行开始的第i列
        for index, value in enumerate(data, start=2):
            sheet.cell(row=index, column=i, value=value)
        # 保存文件
        workbook.save(file_name)
        if len(data) != 0 :
            print(f"关键词-{name}-写入成功!有效数据共{len(data)-data.count(0)}个")

    except Exception as e:
        print(f"发生错误: {e}")


def main(keys,startDate,endDate):
    filename = create_excel(startDate, endDate)
    print(filename+"创建成功！")
    data = []
    i = 2
    for key in keys:
        for year in range(startDate, endDate + 1):
            print(f"正在处理第{year}年，请耐心等待……")
            try:
                file_path = get_index_data(key, year)[0]
                ptbk = get_index_data(key, year)[1]
                res = reCode(file_path,ptbk)
                name = res["name"]
                temp = res["data"]
                data = data + temp
            except:
                continue
        # print(data)
        write_to_excel(filename,name,data,i)
        i = i +1
        data = []
    print("程序运行结束！")


if __name__ == '__main__':
    # 参数列表
    Cookie = 'BIDUPSID=33AD94225F8F2B75579DFF73F1A84EFC; PSTM=1644032496; BAIDUID=EA184A6908BD74FEF96ABCE0319ADDB6:FG=1; BDUSS=JjU1Rldn5YR0FKWElkSGdGWGg5YzZ6TWVyV01GQ3Q2eXEwa1NzMW1jbWJHQUZsSVFBQUFBJCQAAAAAAAAAAAEAAADi0fRewejQocztAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJuL2WSbi9lkbl; H_WISE_SIDS_BFESS=39636; MCITY=-224%3A; Hm_lvt_d101ea4d2a5c67dab98251f0b5de24dc=1703587992; H_PS_PSSID=39996_40043_40052; H_WISE_SIDS=39996_40043_40052; BDORZ=B490B5EBF6F3CD402E515D22BCDA1598; BAIDUID_BFESS=EA184A6908BD74FEF96ABCE0319ADDB6:FG=1; PSINO=7; BA_HECTOR=21a12k0ka425018k018g2g8k8bqipe1ipt1ph1t; ZFY=I7HTSbFydAcJUVmjPzXuzvGXFR8VdngbL:AQCymFmr8k:C; bdindexid=0casdg220ju0881et3mpt69vo6; SIGNIN_UC=70a2711cf1d3d9b1a82d2f87d633bd8a04549005799lPSDPVLvRk%2FeMWzg4YuJMM2zlK1BUMQC%2FRMlU4vfyQbzm6WS0JHMuSSXtOqcywwKfgivmaJY3VF8huVcgy1FB8Yz6w%2B2nA4B6bdoJxMek82RaFlnld4s%2F%2BgaEtC%2Br2JKoxUAICY0hFT7KvsHkjluhtJS2S4sDXzRuLNSXTktF%2FPMg79KkksLnThHIT2Q4sALeXGZfzEU4MiPKvfzn5JAqZ34ctB5gEPdzuwv17VvbLLgIYEAucGtbnR8E9hzAPMo5ZqKSnITOBGI6VDX4ORIHbpC41gn63nE1L95V1zwhSI%3D27433777849423696158428961039005; __cas__rn__=454900579; __cas__st__212=60982c04231e4e2bda92a30acaa5aac80ccd25f1e0c179486857bf01eaf7b04161fd34369d2754291ba4a7b8; __cas__id__212=51122973; CPTK_212=2045301991; CPID_212=51122973; Hm_lpvt_d101ea4d2a5c67dab98251f0b5de24dc=1704888137; ab_sr=1.0.1_NGI1MTZjMTgzOWFlYTY2MzdmMWZkMGIxODI5MDVlODFiM2Y5Zjk1MWMzZDdiN2NiZmM2YWYwNDg2ZTY1MWRiYzE3MWNjNTA2ZGM4NTdmY2E3NzI1MGRkMGVjMzkwYmQ4Nzg1ODM3ZjQ0ZGU4OTI3YTJjYTY3MDQwOTY2Y2YzODkzMzRjNTA1YTE5YTU1MTJlMzAyNWQ3ODM5ZDkyYThmNQ==; BDUSS_BFESS=JjU1Rldn5YR0FKWElkSGdGWGg5YzZ6TWVyV01GQ3Q2eXEwa1NzMW1jbWJHQUZsSVFBQUFBJCQAAAAAAAAAAAEAAADi0fRewejQocztAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAJuL2WSbi9lkbl; RT="z=1&dm=baidu.com&si=62d28216-c3cb-419c-a36d-7195d890f82b&ss=lr7qd20j&sl=2&tt=6mw&bcn=https%3A%2F%2Ffclog.baidu.com%2Flog%2Fweirwood%3Ftype%3Dperf&ld=m7c&ul=1qo2'
    keys = ["人工智能","大数据"]
    # 获取的时间区间，若只获取某一年份，则二者相同
    # 注意！年份区间下限为2011年，不建议选择太早年份
    startDate = 2018
    endDate = 2023
    # 要搜索的关键词，可以输入一个列表
    if Cookie == "":
        Cookie = input("请输入你的Cookie，若错误则无法运行：")
    elif startDate < 2011:
        print("请注意初始年份限制！！！")
    else:
        main(keys,startDate,endDate)


